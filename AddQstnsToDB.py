import MySQLdb
import openpyxl  # For reading Excel files

loc = "./Resources/Questions.xlsx"

# Load the Excel file using openpyxl
wb = openpyxl.load_workbook(loc)
sheet = wb.active  # Access the first (active) worksheet
n = sheet.max_row  # Get the total number of rows

# Establish MySQL database connection
conn = MySQLdb.connect(host='localhost', database='world', user='root', password='root')
cursor = conn.cursor()

# Drop the table if it already exists and create a new one
cursor.execute("DROP TABLE IF EXISTS Questions")
create_table_query = """
CREATE TABLE Questions (
    QID INT,
    qstn TEXT,
    opA TEXT,
    opB TEXT,
    opC TEXT,
    opD TEXT,
    ans INT
)
"""
cursor.execute(create_table_query)

# Code to insert data into the database
for i in range(2, n + 1):  # Loop through rows starting from row 2 to skip header
    try:
        # Extract row data using openpyxl's cell(row, col) method
        lst = [
            sheet.cell(row=i, column=j).value for j in range(1, 8)  # Read columns 1 to 7
        ]
        print(f"Processing row {i}: {lst}")

        # Check if the row has enough data to process
        if None in lst:
            print(f"Skipping row {i} because of incomplete/invalid data")
            continue

        # Parameterized SQL query
        insert_query = """
        INSERT INTO Questions (QID, qstn, opA, opB, opC, opD, ans) 
        VALUES (%s, %s, %s, %s, %s, %s, %s)
        """
        # Prepare data tuple for SQL query
        arg = (
            int(lst[0]),  # Question ID
            lst[1],  # Question text
            lst[2],  # Option A
            lst[3],  # Option B
            lst[4],  # Option C
            lst[5],  # Option D
            int(lst[6])  # Correct answer
        )
        
        # Execute SQL query with parameters
        cursor.execute(insert_query, arg)
        conn.commit()

    except Exception as e:
        print(f"Error processing row {i}: {e}")

# Cleanup database connection
cursor.close()
conn.close()
print("Database connection closed.")
